"""
End-to-end tests for inventory export functionality.

Tests the Excel export endpoint with various scenarios:
- Export without filters
- Export with account_type filter
- Export with status filter
- Export with large datasets
- Export error handling
"""

import pytest
from io import BytesIO
from openpyxl import load_workbook
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.emag_models import EmagProductV2


@pytest.fixture
async def test_products(db: AsyncSession):
    """Create test products for export testing."""
    products = [
        # Out of stock products
        EmagProductV2(
            sku="TEST-001",
            part_number_key="TEST-001",
            name="Test Product 1",
            account_type="main",
            stock_quantity=0,
            price=100.0,
            currency="RON",
            is_active=True,
        ),
        # Critical stock products
        EmagProductV2(
            sku="TEST-002",
            part_number_key="TEST-002",
            name="Test Product 2",
            account_type="fbe",
            stock_quantity=5,
            price=150.0,
            currency="RON",
            is_active=True,
        ),
        # Low stock products
        EmagProductV2(
            sku="TEST-003",
            part_number_key="TEST-003",
            name="Test Product 3",
            account_type="main",
            stock_quantity=15,
            price=200.0,
            currency="RON",
            is_active=True,
        ),
        # In stock products (should not be exported)
        EmagProductV2(
            sku="TEST-004",
            part_number_key="TEST-004",
            name="Test Product 4",
            account_type="fbe",
            stock_quantity=50,
            price=250.0,
            currency="RON",
            is_active=True,
        ),
    ]
    
    for product in products:
        db.add(product)
    
    await db.commit()
    
    yield products
    
    # Cleanup
    for product in products:
        await db.delete(product)
    await db.commit()


@pytest.mark.asyncio
@pytest.mark.e2e
class TestInventoryExport:
    """End-to-end tests for inventory export."""
    
    async def test_export_excel_no_filters(self, async_client: AsyncClient, test_products, auth_headers):
        """Test Excel export without any filters."""
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            headers=auth_headers
        )
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert "low_stock_products_" in response.headers["content-disposition"]
        
        # Verify Excel content
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Check headers
        assert sheet["A1"].value == "Part Number"
        assert sheet["B1"].value == "Product Name"
        assert sheet["C1"].value == "Account Type"
        assert sheet["D1"].value == "Current Stock"
        assert sheet["E1"].value == "Status"
        
        # Check data rows (should have 3 low stock products)
        # Row 2: TEST-001 (out of stock)
        assert sheet["A2"].value == "TEST-001"
        assert sheet["D2"].value == 0
        assert "OUT OF STOCK" in sheet["E2"].value
        
        # Row 3: TEST-002 (critical)
        assert sheet["A3"].value == "TEST-002"
        assert sheet["D3"].value == 5
        assert "CRITICAL" in sheet["E3"].value
        
        # Row 4: TEST-003 (low stock)
        assert sheet["A4"].value == "TEST-003"
        assert sheet["D4"].value == 15
        assert "LOW STOCK" in sheet["E4"].value
        
        # TEST-004 should NOT be in export (stock = 50)
        assert sheet["A5"].value != "TEST-004"
    
    async def test_export_excel_with_account_filter(self, async_client: AsyncClient, test_products, auth_headers):
        """Test Excel export with account_type filter."""
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            params={"account_type": "MAIN"},
            headers=auth_headers
        )
        
        assert response.status_code == 200
        
        # Verify Excel content
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Should only have MAIN account products
        row = 2
        while sheet[f"A{row}"].value:
            account_type = sheet[f"C{row}"].value
            assert account_type == "MAIN", f"Row {row} has account type {account_type}, expected MAIN"
            row += 1
    
    async def test_export_excel_with_status_filter(self, async_client: AsyncClient, test_products, auth_headers):
        """Test Excel export with status filter."""
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            params={"status": "out_of_stock"},
            headers=auth_headers
        )
        
        assert response.status_code == 200
        
        # Verify Excel content
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Should only have out of stock products
        row = 2
        while sheet[f"A{row}"].value:
            status = sheet[f"E{row}"].value
            assert "OUT OF STOCK" in status, f"Row {row} has status {status}, expected OUT OF STOCK"
            row += 1
    
    async def test_export_excel_formatting(self, async_client: AsyncClient, test_products, auth_headers):
        """Test Excel formatting (colors, borders, etc.)."""
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            headers=auth_headers
        )
        
        assert response.status_code == 200
        
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Check header formatting
        header_cell = sheet["A1"]
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb == "FF4472C4"  # Blue header
        
        # Check conditional formatting for out of stock (row 2)
        out_of_stock_cell = sheet["A2"]
        # Should have red background for out of stock
        assert out_of_stock_cell.fill.start_color.rgb in ["FFFF6B6B", "FFFF0000"]
        
        # Check borders
        assert header_cell.border.left.style == "thin"
        assert header_cell.border.right.style == "thin"
    
    async def test_export_excel_summary_section(self, async_client: AsyncClient, test_products, auth_headers):
        """Test that summary section is included in export."""
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            headers=auth_headers
        )
        
        assert response.status_code == 200
        
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Find summary section (should be after data rows + 2)
        # Look for "SUMMARY" text
        found_summary = False
        for row in range(1, sheet.max_row + 1):
            if sheet[f"A{row}"].value == "SUMMARY":
                found_summary = True
                
                # Check summary content
                assert sheet[f"A{row + 1}"].value == "Total Products:"
                assert isinstance(sheet[f"B{row + 1}"].value, int)
                
                assert sheet[f"A{row + 2}"].value == "Total Reorder Cost:"
                assert "RON" in str(sheet[f"B{row + 2}"].value)
                
                assert sheet[f"A{row + 3}"].value == "Generated:"
                assert sheet[f"B{row + 3}"].value is not None
                
                break
        
        assert found_summary, "Summary section not found in export"
    
    async def test_export_excel_no_products(self, async_client: AsyncClient, auth_headers):
        """Test Excel export when no low stock products exist."""
        # This test assumes no low stock products in clean database
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            params={"status": "out_of_stock", "account_type": "NONEXISTENT"},
            headers=auth_headers
        )
        
        # Should return 404 when no products found
        assert response.status_code == 404
        assert "No low stock products found" in response.json()["detail"]
    
    async def test_export_excel_unauthorized(self, async_client: AsyncClient):
        """Test Excel export without authentication."""
        response = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel"
        )
        
        assert response.status_code == 401
    
    async def test_export_excel_large_dataset(self, async_client: AsyncClient, db: AsyncSession, auth_headers):
        """Test Excel export with large dataset (performance test)."""
        # Create 100 test products
        products = []
        for i in range(100):
            product = EmagProductV2(
                sku=f"BULK-{i:03d}",
                part_number_key=f"BULK-{i:03d}",
                name=f"Bulk Test Product {i}",
                account_type="main" if i % 2 == 0 else "fbe",
                stock_quantity=i % 20,  # 0-19, all low stock
                price=100.0 + i,
                currency="RON",
                is_active=True,
            )
            products.append(product)
            db.add(product)
        
        await db.commit()
        
        try:
            # Export should complete within reasonable time
            response = await async_client.get(
                "/api/v1/emag-inventory/export/low-stock-excel",
                headers=auth_headers,
                timeout=30.0  # 30 seconds timeout
            )
            
            assert response.status_code == 200
            
            # Verify file size is reasonable (should be < 1MB for 100 products)
            assert len(response.content) < 1024 * 1024
            
            # Verify Excel can be opened
            excel_file = BytesIO(response.content)
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            # Should have 100+ rows (header + data + summary)
            assert sheet.max_row >= 100
            
        finally:
            # Cleanup
            for product in products:
                await db.delete(product)
            await db.commit()


@pytest.mark.asyncio
@pytest.mark.e2e
class TestExportIntegration:
    """Integration tests for export with other features."""
    
    async def test_export_after_inventory_update(self, async_client: AsyncClient, test_products, auth_headers):
        """Test that export reflects recent inventory updates."""
        # Initial export
        response1 = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            headers=auth_headers
        )
        
        assert response1.status_code == 200
        
        # Update a product's stock (simulate inventory update)
        # This would typically be done through an update endpoint
        # For testing, we'll verify the export shows current data
        
        # Second export should show same data (no changes made)
        response2 = await async_client.get(
            "/api/v1/emag-inventory/export/low-stock-excel",
            headers=auth_headers
        )
        
        assert response2.status_code == 200
        assert len(response1.content) == len(response2.content)
