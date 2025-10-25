"""
Low Stock Products with Supplier Selection API

This endpoint provides low stock products with all available suppliers,
allowing users to select suppliers and export filtered data.
"""

import logging
from datetime import datetime, timedelta
from io import BytesIO

import requests
from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from PIL import Image
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.db import get_db
from app.models.emag_models import EmagOrder, EmagProductV2
from app.models.inventory import InventoryItem, Warehouse
from app.models.order import OrderLine
from app.models.product import Product
from app.models.product_supplier_sheet import ProductSupplierSheet
from app.models.purchase import PurchaseOrder, PurchaseOrderItem
from app.models.sales import SalesOrderLine
from app.models.supplier import SupplierProduct
from app.models.user import User
from app.security.jwt import get_current_user

router = APIRouter(prefix="/inventory", tags=["low-stock-suppliers"])


# ============================================================================
# Helper Functions
# ============================================================================


def calculate_stock_status(item: InventoryItem) -> str:
    """Calculate stock status based on quantity and thresholds."""
    available = item.quantity - item.reserved_quantity

    if available <= 0:
        return "out_of_stock"
    elif available <= item.minimum_stock:
        return "critical"
    elif available <= item.reorder_point:
        return "low_stock"
    else:
        return "in_stock"


def calculate_reorder_quantity(item: InventoryItem) -> int:
    """
    Calculate recommended reorder quantity.

    Priority:
    1. Manual override (manual_reorder_quantity) - if set, use this value
    2. Automatic calculation based on stock levels
    """
    # If manual reorder quantity is set, use it (override automatic calculation)
    if item.manual_reorder_quantity is not None:
        return item.manual_reorder_quantity

    # Otherwise, calculate automatically
    available = item.quantity - item.reserved_quantity

    if item.maximum_stock:
        # Reorder to maximum stock
        return max(0, item.maximum_stock - available)
    elif item.reorder_point > 0:
        # Reorder to double the reorder point
        return max(0, (item.reorder_point * 2) - available)
    else:
        # Default: reorder to minimum stock * 3
        return max(0, (item.minimum_stock * 3) - available)


async def calculate_sold_quantity_last_6_months(
    db: AsyncSession, product_ids: list[int]
) -> dict[int, dict]:
    """
    Calculate quantity sold in the last 6 months for each product.

    Aggregates data from:
    1. eMAG Orders (from products JSONB field)
    2. Sales Orders (SalesOrderLine)
    3. Generic Orders (OrderLine)

    Returns:
        dict: {product_id: {"total_sold": int, "avg_monthly": float, "sources": dict}}
    """
    if not product_ids:
        return {}

    six_months_ago = datetime.now() - timedelta(days=180)
    # Initialize each product with its own dict
    sold_data = {
        pid: {"total_sold": 0, "avg_monthly": 0.0, "sources": {}}
        for pid in product_ids
    }

    # Get product SKUs mapping
    # CRITICAL: We need to map both Product.sku AND emag part_number_key to product_id
    # because emag_orders uses part_number_key (e.g., DVX0FSYBM) not Product.sku (e.g., EMG463)
    products_query = select(Product.id, Product.sku).where(Product.id.in_(product_ids))
    products_result = await db.execute(products_query)
    product_sku_map = {pid: sku for pid, sku in products_result.all()}
    sku_to_id_map = {sku: pid for pid, sku in product_sku_map.items()}

    # CRITICAL FIX: Also map emag part_number_key to product_id
    # Query emag_products_v2 to get part_number_key for our products
    try:
        from app.models.emag_models import EmagProductV2
        emag_products_query = (
            select(EmagProductV2.sku, EmagProductV2.part_number_key)
            .where(EmagProductV2.sku.in_(list(product_sku_map.values())))
        )
        emag_products_result = await db.execute(emag_products_query)

        # Map part_number_key -> product_id
        for local_sku, part_number_key in emag_products_result.all():
            if part_number_key and local_sku in sku_to_id_map:
                product_id = sku_to_id_map[local_sku]
                # Add part_number_key mapping
                sku_to_id_map[part_number_key] = product_id
                logging.debug(
                    "Mapped eMAG part_number_key %s -> product_id %s (local SKU: %s)",
                    part_number_key,
                    product_id,
                    local_sku,
                )
    except Exception as e:
        logging.warning(f"Error mapping eMAG part_number_keys: {e}")

    # 1. Query eMAG Orders (products stored in JSONB)
    # We need to extract product quantities from the JSONB products field
    # NOTE: Table may not exist in all environments - handle gracefully
    try:
        emag_orders_query = (
            select(EmagOrder.products, EmagOrder.order_date)
            .where(
                and_(
                    EmagOrder.order_date >= six_months_ago,
                    EmagOrder.status.in_([3, 4]),  # 3=prepared, 4=finalized
                    EmagOrder.products.isnot(None),
                )
            )
        )
        emag_result = await db.execute(emag_orders_query)
        emag_orders = emag_result.all()

        # Process eMAG orders
        for products_json, _ in emag_orders:
            if not products_json or not isinstance(products_json, list):
                continue

            for product_item in products_json:
                # eMAG products structure: {"part_number_key": "SKU", "quantity": 1, ...}
                sku = product_item.get("part_number_key") or product_item.get("sku")
                quantity = product_item.get("quantity", 0)

                if sku and sku in sku_to_id_map:
                    product_id = sku_to_id_map[sku]
                    sold_data[product_id]["total_sold"] += quantity
                    sold_data[product_id]["sources"]["emag"] = (
                        sold_data[product_id]["sources"].get("emag", 0) + quantity
                    )
    except Exception as e:
        logging.warning(f"Error querying eMAG orders (table may not exist): {e}")

    # 2. Query Sales Orders
    sales_query = (
        select(SalesOrderLine.product_id, func.sum(SalesOrderLine.quantity))
        .join(
            SalesOrderLine.sales_order
        )
        .where(
            and_(
                SalesOrderLine.product_id.in_(product_ids),
                SalesOrderLine.sales_order.has(
                    and_(
                        func.date(
                            SalesOrderLine.sales_order.property.mapper.class_.order_date
                        )
                        >= six_months_ago,
                        SalesOrderLine.sales_order.property.mapper.class_.status.in_(
                            ["confirmed", "processing", "shipped", "delivered"]
                        ),
                    )
                ),
            )
        )
        .group_by(SalesOrderLine.product_id)
    )

    try:
        sales_result = await db.execute(sales_query)
        for product_id, quantity in sales_result.all():
            if product_id in sold_data:
                qty = int(quantity or 0)
                sold_data[product_id]["total_sold"] += qty
                sold_data[product_id]["sources"]["sales_orders"] = qty
    except Exception as e:
        logging.warning(f"Error querying sales orders: {e}")

    # 3. Query Generic Orders
    orders_query = (
        select(OrderLine.product_id, func.sum(OrderLine.quantity))
        .join(OrderLine.order)
        .where(
            and_(
                OrderLine.product_id.in_(product_ids),
                OrderLine.order.has(
                    and_(
                        func.date(
                            OrderLine.order.property.mapper.class_.order_date
                        )
                        >= six_months_ago,
                        OrderLine.order.property.mapper.class_.status.in_(
                            [
                                "confirmed",
                                "processing",
                                "shipped",
                                "delivered",
                                "completed",
                            ]
                        ),
                    )
                ),
            )
        )
        .group_by(OrderLine.product_id)
    )

    try:
        orders_result = await db.execute(orders_query)
        for product_id, quantity in orders_result.all():
            if product_id in sold_data:
                qty = int(quantity or 0)
                sold_data[product_id]["total_sold"] += qty
                sold_data[product_id]["sources"]["orders"] = qty
    except Exception as e:
        logging.warning(f"Error querying generic orders: {e}")

    # Calculate average monthly sales
    for product_id in sold_data:
        total = sold_data[product_id]["total_sold"]
        sold_data[product_id]["avg_monthly"] = round(total / 6.0, 2)

    return sold_data


# ============================================================================
# API Endpoints
# ============================================================================


@router.get("/low-stock-with-suppliers")
async def get_low_stock_with_suppliers(
    warehouse_id: int | None = Query(None, description="Filter by warehouse"),
    account_type: str | None = Query(
        None, description="Filter by account type: main or fbe (maps to warehouse)"
    ),
    status: str | None = Query(
        None, description="Filter by status: critical, low_stock, out_of_stock"
    ),
    include_inactive: bool = Query(False, description="Include inactive products"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get low stock products with all available suppliers.

    Args:
        warehouse_id: Direct warehouse filter (takes precedence)
        account_type: Filter by account (main/fbe) - maps to warehouse codes
        status: Stock status filter
        include_inactive: Include inactive products
        skip: Pagination offset
        limit: Maximum results

    Returns:
    - Product details with current stock
    - List of all suppliers for each product with prices
    - Allows frontend to select suppliers for export
    """

    # Map account_type to warehouse codes and eMAG account type
    # If account_type is provided, find matching warehouses
    warehouse_codes = None
    emag_account_type = None
    if account_type:
        account_type_lower = account_type.lower().strip()
        if account_type_lower == "fbe":
            warehouse_codes = ["EMAG-FBE", "FBE", "eMAG-FBE"]
            emag_account_type = "fbe"
        elif account_type_lower == "main":
            warehouse_codes = ["MAIN", "EMAG-MAIN", "eMAG-MAIN", "PRIMARY"]
            emag_account_type = "main"

    # Build query for low stock products with eMAG PNK
    # Use LEFT JOIN with account_type filter to avoid duplicates
    emag_join_condition = Product.sku == EmagProductV2.sku
    if emag_account_type:
        emag_join_condition = and_(
            Product.sku == EmagProductV2.sku,
            EmagProductV2.account_type == emag_account_type,
        )

    query = (
        select(InventoryItem, Product, Warehouse, EmagProductV2)
        .join(Product, InventoryItem.product_id == Product.id)
        .join(Warehouse, InventoryItem.warehouse_id == Warehouse.id)
        .outerjoin(EmagProductV2, emag_join_condition)
        .where(InventoryItem.is_active.is_(True))
    )

    # Filter by warehouse (direct filter takes precedence)
    if warehouse_id:
        query = query.where(InventoryItem.warehouse_id == warehouse_id)
    elif warehouse_codes:
        # Filter by warehouse codes based on account_type
        query = query.where(Warehouse.code.in_(warehouse_codes))

    # Filter by product active status
    if not include_inactive:
        query = query.where(Product.is_active.is_(True))

    # Filter by stock status
    if status == "out_of_stock":
        query = query.where(InventoryItem.quantity <= 0)
    elif status == "critical":
        query = query.where(
            and_(
                InventoryItem.quantity > 0,
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                <= InventoryItem.minimum_stock,
            )
        )
    elif status == "low_stock":
        query = query.where(
            and_(
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                > InventoryItem.minimum_stock,
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                <= InventoryItem.reorder_point,
            )
        )
    else:
        # Default: show all products that need attention
        query = query.where(
            or_(
                InventoryItem.quantity <= 0,
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                <= InventoryItem.reorder_point,
            )
        )

    # Order by urgency
    query = query.order_by(
        InventoryItem.quantity.asc(),
        (InventoryItem.quantity - InventoryItem.reserved_quantity).asc(),
    )

    # Get total count
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Apply pagination
    query = query.offset(skip).limit(limit)

    # Execute query
    result = await db.execute(query)
    items = result.all()

    # Get all product IDs for supplier lookup
    product_ids = [item[1].id for item in items]

    # Calculate sold quantities for last 6 months
    sold_quantities = await calculate_sold_quantity_last_6_months(db, product_ids)

    # Get pending purchase orders for these products
    pending_orders_query = (
        select(PurchaseOrderItem, PurchaseOrder)
        .join(PurchaseOrder, PurchaseOrderItem.purchase_order_id == PurchaseOrder.id)
        .where(
            and_(
                PurchaseOrderItem.local_product_id.in_(product_ids),
                PurchaseOrder.status.in_(["sent", "confirmed", "partially_received"]),
                (PurchaseOrderItem.quantity_received or 0) < PurchaseOrderItem.quantity_ordered,
            )
        )
    )
    pending_orders_result = await db.execute(pending_orders_query)
    pending_orders = pending_orders_result.all()

    # Organize pending orders by product
    pending_by_product = {}
    for line, po in pending_orders:
        product_id = line.local_product_id  # Use local_product_id from PurchaseOrderItem
        if product_id not in pending_by_product:
            pending_by_product[product_id] = []

        pending_qty = line.quantity_ordered - (line.quantity_received or 0)
        pending_by_product[product_id].append({
            "purchase_order_id": po.id,
            "order_number": po.order_number,
            "ordered_quantity": line.quantity_ordered,
            "received_quantity": line.quantity_received or 0,
            "pending_quantity": pending_qty,
            "expected_delivery_date": po.expected_delivery_date.isoformat()
            if po.expected_delivery_date
            else None,
            "status": po.status,
        })

    # Get suppliers from ProductSupplierSheet (Google Sheets data)
    supplier_sheets_query = (
        select(ProductSupplierSheet)
        .join(Product, ProductSupplierSheet.sku == Product.sku)
        .where(
            and_(Product.id.in_(product_ids), ProductSupplierSheet.is_active.is_(True))
        )
    )
    supplier_sheets_result = await db.execute(supplier_sheets_query)
    supplier_sheets = supplier_sheets_result.scalars().all()

    # Get suppliers from SupplierProduct (1688.com data)
    supplier_products_query = (
        select(SupplierProduct)
        .options(selectinload(SupplierProduct.supplier))
        .where(
            and_(
                SupplierProduct.local_product_id.in_(product_ids),
                SupplierProduct.is_active.is_(True),
            )
        )
    )
    supplier_products_result = await db.execute(supplier_products_query)
    supplier_products = supplier_products_result.scalars().all()

    # Organize suppliers by product with deduplication
    suppliers_by_product = {}

    # Track seen suppliers by URL to avoid duplicates
    seen_suppliers_by_product = {}  # product_id -> set of URLs

    # PRIORITY CHANGE: Add 1688.com suppliers FIRST (these are the primary source)
    for sp in supplier_products:
        if sp.local_product_id not in suppliers_by_product:
            suppliers_by_product[sp.local_product_id] = []
            seen_suppliers_by_product[sp.local_product_id] = set()

        # Create unique key for deduplication
        supplier_url = sp.supplier_product_url or ""
        dedup_key = (
            supplier_url.strip().lower()
            if supplier_url
            else f"{sp.supplier.name if sp.supplier else 'Unknown'}_{sp.supplier_price}"
        )

        # Skip if already seen (duplicate from Google Sheets)
        if dedup_key in seen_suppliers_by_product[sp.local_product_id]:
            continue

        seen_suppliers_by_product[sp.local_product_id].add(dedup_key)

        suppliers_by_product[sp.local_product_id].append(
            {
                "supplier_id": f"1688_{sp.id}",
                "actual_supplier_id": sp.supplier_id,  # Real supplier ID for API calls
                "supplier_product_id": sp.id,  # SupplierProduct ID for updates
                "supplier_name": sp.supplier.name if sp.supplier else "Unknown",
                "supplier_type": "1688",
                "price": sp.supplier_price,
                "currency": sp.supplier_currency,
                "price_ron": None,  # Could be calculated if exchange rate available
                "supplier_url": sp.supplier_product_url,
                "supplier_contact": sp.supplier.email if sp.supplier else None,
                "chinese_name": sp.supplier_product_chinese_name
                or sp.supplier_product_name,
                "specification": sp.supplier_product_specification,
                "is_preferred": sp.is_preferred,
                "is_verified": sp.manual_confirmed,
                "last_updated": sp.last_price_update.isoformat()
                if sp.last_price_update
                else None,
            }
        )

    # Add Google Sheets suppliers (only if not already present from 1688)
    for sheet in supplier_sheets:
        # Find product by SKU
        product_id = next(
            (item[1].id for item in items if item[1].sku == sheet.sku), None
        )
        if product_id:
            if product_id not in suppliers_by_product:
                suppliers_by_product[product_id] = []
                seen_suppliers_by_product[product_id] = set()

            # Create unique key for deduplication (URL is most reliable)
            supplier_url = sheet.supplier_url or ""
            dedup_key = (
                supplier_url.strip().lower()
                if supplier_url
                else f"{sheet.supplier_name}_{sheet.price_cny}"
            )

            # Skip if already seen (duplicate from 1688)
            if dedup_key in seen_suppliers_by_product[product_id]:
                continue

            seen_suppliers_by_product[product_id].add(dedup_key)

            suppliers_by_product[product_id].append(
                {
                    "supplier_id": f"sheet_{sheet.id}",
                    "sheet_id": sheet.id,  # Sheet ID for updates
                    "supplier_name": sheet.supplier_name,
                    "supplier_type": "google_sheets",
                    "price": sheet.price_cny,
                    "currency": "CNY",
                    "price_ron": sheet.calculated_price_ron,
                    "supplier_url": sheet.supplier_url,
                    "supplier_contact": sheet.supplier_contact,
                    "chinese_name": sheet.supplier_product_chinese_name,
                    "specification": sheet.supplier_product_specification,
                    "is_preferred": sheet.is_preferred,
                    "is_verified": sheet.is_verified,
                    "last_updated": sheet.price_updated_at.isoformat()
                    if sheet.price_updated_at
                    else None,
                }
            )

    # Format response
    products_data = []
    for inventory_item, product, warehouse, emag_product in items:
        available_qty = inventory_item.quantity - inventory_item.reserved_quantity
        stock_status = calculate_stock_status(inventory_item)
        reorder_qty = calculate_reorder_quantity(inventory_item)

        # Get suppliers for this product
        suppliers = suppliers_by_product.get(product.id, [])

        # Sort suppliers: preferred first, then by price
        suppliers.sort(
            key=lambda s: (not s["is_preferred"], s["price"] or float("inf"))
        )

        # Get PNK and URL from eMAG product if available, otherwise from product table
        part_number_key = (
            emag_product.part_number_key
            if emag_product
            else product.emag_part_number_key
        )

        # Get product URL from eMAG (seller website URL)
        product_url = emag_product.url if emag_product else None

        # Get pending orders for this product
        pending_orders = pending_by_product.get(product.id, [])
        total_pending_quantity = sum(order["pending_quantity"] for order in pending_orders)

        # Calculate adjusted reorder quantity (subtract pending orders)
        adjusted_reorder_qty = max(0, reorder_qty - total_pending_quantity)

        # Get sold quantity data
        sold_data = sold_quantities.get(product.id, {
            "total_sold": 0,
            "avg_monthly": 0.0,
            "sources": {}
        })

        products_data.append(
            {
                "inventory_item_id": inventory_item.id,
                "product_id": product.id,
                "sku": product.sku,
                "name": product.name,
                "chinese_name": product.chinese_name,
                "part_number_key": part_number_key,
                "product_url": product_url,
                "image_url": product.image_url,
                "warehouse_id": warehouse.id,
                "warehouse_name": warehouse.name,
                "warehouse_code": warehouse.code,
                "quantity": inventory_item.quantity,
                "reserved_quantity": inventory_item.reserved_quantity,
                "available_quantity": available_qty,
                "minimum_stock": inventory_item.minimum_stock,
                "reorder_point": inventory_item.reorder_point,
                "maximum_stock": inventory_item.maximum_stock,
                "manual_reorder_quantity": inventory_item.manual_reorder_quantity,
                "unit_cost": inventory_item.unit_cost,
                "stock_status": stock_status,
                "reorder_quantity": reorder_qty,
                "adjusted_reorder_quantity": adjusted_reorder_qty,
                "location": inventory_item.location,
                "base_price": product.base_price,
                "currency": product.currency,
                "is_discontinued": product.is_discontinued,
                "suppliers": suppliers,
                "supplier_count": len(suppliers),
                "pending_orders": pending_orders,
                "total_pending_quantity": total_pending_quantity,
                "has_pending_orders": len(pending_orders) > 0,
                "sold_last_6_months": sold_data["total_sold"],
                "avg_monthly_sales": sold_data["avg_monthly"],
                "sales_sources": sold_data["sources"],
            }
        )

    return {
        "status": "success",
        "data": {
            "products": products_data,
            "pagination": {
                "total": total,
                "skip": skip,
                "limit": limit,
                "has_more": skip + limit < total,
            },
            "summary": {
                "total_low_stock": total,
                "out_of_stock": sum(
                    1 for p in products_data if p["stock_status"] == "out_of_stock"
                ),
                "critical": sum(
                    1 for p in products_data if p["stock_status"] == "critical"
                ),
                "low_stock": sum(
                    1 for p in products_data if p["stock_status"] == "low_stock"
                ),
                "products_with_suppliers": sum(
                    1 for p in products_data if p["supplier_count"] > 0
                ),
                "products_without_suppliers": sum(
                    1 for p in products_data if p["supplier_count"] == 0
                ),
                "products_with_pending_orders": sum(
                    1 for p in products_data if p["has_pending_orders"]
                ),
                "total_pending_quantity": sum(
                    p["total_pending_quantity"] for p in products_data
                ),
            },
        },
    }


@router.post("/export/low-stock-by-supplier")
async def export_low_stock_by_supplier(
    selected_products: list[dict] = Body(
        ..., description="List of products with selected supplier IDs"
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Export low stock products grouped by selected suppliers to Excel.

    Request body format:
    [
        {
            "product_id": 123,
            "sku": "PROD-001",
            "supplier_id": "sheet_45",
            "reorder_quantity": 50
        },
        ...
    ]

    Creates separate sheets for each supplier in the Excel file.
    """

    # Validation: Check Excel library
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl",
        )

    # Validation: Check if products selected
    if not selected_products:
        raise HTTPException(status_code=400, detail="No products selected for export")

    # Validation: Check reasonable limit (prevent abuse)
    if len(selected_products) > 1000:
        raise HTTPException(
            status_code=400,
            detail="Too many products selected. Maximum 1000 products per export",
        )

    # First pass: collect all supplier IDs and map to names
    supplier_id_to_name = {}
    for item in selected_products:
        supplier_id = item.get("supplier_id")
        supplier_name = item.get("supplier_name", "Unknown")
        if supplier_id:
            supplier_id_to_name[supplier_id] = supplier_name

    # Group products by supplier NAME (not ID) to merge same suppliers from different sources
    products_by_supplier_name = {}
    for item in selected_products:
        supplier_id = item.get("supplier_id")
        supplier_name = item.get("supplier_name", "Unknown")

        if supplier_name:
            if supplier_name not in products_by_supplier_name:
                products_by_supplier_name[supplier_name] = {
                    "products": [],
                    "supplier_ids": [],  # Track all IDs for this supplier
                }
            products_by_supplier_name[supplier_name]["products"].append(item)
            if (
                supplier_id
                not in products_by_supplier_name[supplier_name]["supplier_ids"]
            ):
                products_by_supplier_name[supplier_name]["supplier_ids"].append(
                    supplier_id
                )

    # Get product and supplier details from database
    product_ids = [item["product_id"] for item in selected_products]

    # Get products with inventory
    products_query = (
        select(Product, InventoryItem, Warehouse)
        .join(InventoryItem, Product.id == InventoryItem.product_id)
        .join(Warehouse, InventoryItem.warehouse_id == Warehouse.id)
        .where(Product.id.in_(product_ids))
    )
    products_result = await db.execute(products_query)
    products_data = {p.id: (p, inv, wh) for p, inv, wh in products_result.all()}

    # Get all supplier sheet data
    supplier_sheets_query = select(ProductSupplierSheet)
    supplier_sheets_result = await db.execute(supplier_sheets_query)
    supplier_sheets = {
        f"sheet_{s.id}": s for s in supplier_sheets_result.scalars().all()
    }

    # Get all 1688 supplier data
    supplier_products_query = select(SupplierProduct).options(
        selectinload(SupplierProduct.supplier)
    )
    supplier_products_result = await db.execute(supplier_products_query)
    supplier_products_map = {
        f"1688_{sp.id}": sp for sp in supplier_products_result.scalars().all()
    }

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_font = Font(bold=False, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Cell alignment styles
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    critical_fill = PatternFill(
        start_color="ffffff", end_color="ffffff", fill_type="solid"
    )
    low_stock_fill = PatternFill(
        start_color="ffffff", end_color="ffffff", fill_type="solid"
    )

    # Create a sheet for each supplier (grouped by name)
    for supplier_name, supplier_data in products_by_supplier_name.items():
        supplier_products_list = supplier_data["products"]
        supplier_ids = supplier_data["supplier_ids"]

        # Get supplier info from first available ID
        supplier_contact = ""
        supplier_type = ""
        supplier_url_primary = ""

        # Try to get info from all supplier IDs (prefer Google Sheets)
        for supplier_id in supplier_ids:
            if supplier_id.startswith("sheet_"):
                sheet_data = supplier_sheets.get(supplier_id)
                if sheet_data:
                    supplier_contact = sheet_data.supplier_contact or supplier_contact
                    supplier_type = (
                        "Google Sheets"
                        if not supplier_type
                        else f"{supplier_type}, Google Sheets"
                    )
                    supplier_url_primary = (
                        sheet_data.supplier_url or supplier_url_primary
                    )
            elif supplier_id.startswith("1688_"):
                sp_data = supplier_products_map.get(supplier_id)
                if sp_data and sp_data.supplier:
                    supplier_contact = (
                        sp_data.supplier.email
                        or sp_data.supplier.phone
                        or supplier_contact
                    )
                    supplier_type = (
                        "1688.com"
                        if not supplier_type
                        else f"{supplier_type}, 1688.com"
                    )
                    supplier_url_primary = (
                        sp_data.supplier_product_url or supplier_url_primary
                    )

        # Create sheet (Excel sheet names max 31 chars)
        sheet_name = (
            supplier_name[:28] + "..." if len(supplier_name) > 31 else supplier_name
        )
        ws = wb.create_sheet(title=sheet_name)

        # Add supplier header
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = "GALACTRONICE ROMANIA S.R.L."
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        ws.merge_cells("A2:H3")
        contact_cell = ws["A2"]
        contact_cell.value = "您能在包裹上贴上公司的标签吗？"
        contact_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "图片",
            "名称",
            "规格名",
            "数量",
            "零售价",
            "金额",
            "商品链接",
            "图片链接",
        ]

        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Add products
        row_num = 5
        total_cost = 0

        for product_item in supplier_products_list:
            product_id = product_item["product_id"]
            reorder_qty = product_item.get("reorder_quantity", 0)
            product_supplier_id = product_item.get("supplier_id")

            if product_id not in products_data:
                continue

            product, inventory, warehouse = products_data[product_id]

            # Get supplier price and specification for THIS specific product's supplier ID
            unit_price = 0
            # currency = "CNY"  # Not used in current implementation
            supplier_url = ""
            specification = ""

            if product_supplier_id and product_supplier_id.startswith("sheet_"):
                sheet_data = supplier_sheets.get(product_supplier_id)
                if sheet_data:
                    unit_price = sheet_data.price_cny
                    supplier_url = sheet_data.supplier_url or ""
                    specification = sheet_data.supplier_product_specification or ""
            elif product_supplier_id and product_supplier_id.startswith("1688_"):
                sp_data = supplier_products_map.get(product_supplier_id)
                if sp_data:
                    unit_price = sp_data.supplier_price
                    # currency = sp_data.supplier_currency or "CNY"
                    # Not used in current implementation
                    supplier_url = sp_data.supplier_product_url or ""
                    specification = sp_data.supplier_product_specification or ""

            total_price = unit_price * reorder_qty
            total_cost += total_price

            stock_status = calculate_stock_status(inventory)

            # Write row
            row_data = [
                "",  # Image placeholder
                product.chinese_name or "",
                specification,
                reorder_qty,
                unit_price,
                total_price,
                supplier_url,
                product.image_url or "",
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

                # Apply alignment based on column
                # Columns:
                #   1=图片, 2=名称, 3=规格名, 4=数量, 5=零售价,
                #   6=金额, 7=商品链接, 8=图片链接
                if col_num in [4, 5, 6, 7, 8]:  # 数量, 零售价, 金额, 商品链接, 图片链接
                    cell.alignment = center_alignment
                else:  # 图片, 名称, 规格名
                    cell.alignment = left_alignment

                # Add hyperlink for 商品链接 column (column 7)
                if col_num == 7 and value:  # 商品链接
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                    cell.alignment = center_alignment
                    cell.border = border

                # Apply status coloring
                if stock_status == "out_of_stock" or stock_status == "critical":
                    cell.fill = critical_fill
                elif stock_status == "low_stock":
                    cell.fill = low_stock_fill

            # Insert product image in first column
            if product.image_url:
                try:
                    # Download image
                    response = requests.get(product.image_url, timeout=5)
                    if response.status_code == 200:
                        img_data = BytesIO(response.content)
                        img = Image.open(img_data)

                        # Resize image to fit in cell (max 80x80 pixels)
                        img.thumbnail((183, 183), Image.Resampling.LANCZOS)

                        # Save to BytesIO
                        img_buffer = BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)

                        # Create Excel image
                        excel_img = ExcelImage(img_buffer)
                        excel_img.width = 183
                        excel_img.height = 183

                        # Add image to cell
                        ws.add_image(excel_img, f"A{row_num}")

                        # Set row height to accommodate image
                        ws.row_dimensions[row_num].height = 139
                except Exception as e:
                    # If image download/processing fails, continue without image
                    logging.warning(f"Failed to process image for product {product.sku}: {e}")

            row_num += 1

        # Add summary
        row_num += 1
        summary_cell = ws.cell(row=row_num, column=1)
        summary_cell.value = "金额:"
        summary_cell.font = Font(bold=False, size=13)

        total_cell = ws.cell(row=row_num, column=6)
        total_cell.value = total_cost
        total_cell.font = Font(bold=True, size=13)
        total_cell.fill = PatternFill(
            start_color="8fce00", end_color="8fce00", fill_type="solid"
        )

        row_num += 1
        ws.cell(
            row=row_num, column=1
        ).value = f"Total Products: {len(supplier_products_list)}"
        ws.cell(row=row_num, column=1).font = Font(bold=True)

        row_num += 1
        ws.cell(
            row=row_num, column=1
        ).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Adjust column widths
        column_widths = [23, 60, 30, 9, 9, 9, 26, 23]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header rows
        ws.freeze_panes = "A5"

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"low_stock_by_supplier_{timestamp}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
