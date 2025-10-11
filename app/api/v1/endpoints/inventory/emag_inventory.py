"""
eMAG Inventory Management Endpoints.

Provides endpoints for monitoring and managing eMAG inventory including:
- Low stock alerts
- Inventory statistics
- Stock level monitoring across MAIN and FBE accounts
"""

from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.core.logging import get_logger
from app.core.utils.account_utils import normalize_account_type
from app.db import get_db
from app.models.emag_models import EmagProductV2
from app.security.jwt import get_current_user

logger = get_logger(__name__)

try:
    from app.services.inventory.inventory_cache_service import get_inventory_cache

    CACHE_AVAILABLE = True
except ImportError:
    CACHE_AVAILABLE = False
    logger.warning("Inventory cache service not available - caching disabled")

router = APIRouter(prefix="/emag-inventory", tags=["emag-inventory"])


# ============================================================================
# Helper Functions
# ============================================================================


def calculate_stock_status(
    stock_quantity: int, min_stock: int = 10, reorder_point: int = 20
) -> str:
    """
    Calculate stock status based on quantity.

    Args:
        stock_quantity: Current stock level
        min_stock: Minimum stock threshold (default: 10)
        reorder_point: Reorder point threshold (default: 20)

    Returns:
        Status string: 'out_of_stock', 'critical', 'low_stock', or 'in_stock'
    """
    if stock_quantity <= 0:
        return "out_of_stock"
    elif stock_quantity <= min_stock:
        return "critical"
    elif stock_quantity <= reorder_point:
        return "low_stock"
    else:
        return "in_stock"


def calculate_reorder_quantity(
    stock_quantity: int, max_stock: int = 100, target_stock: int = 20
) -> int:
    """
    Calculate recommended reorder quantity.

    Args:
        stock_quantity: Current stock level
        max_stock: Maximum stock to order when out of stock (default: 100)
        target_stock: Target stock level (default: 20)

    Returns:
        Recommended quantity to reorder
    """
    if stock_quantity <= 0:
        return max_stock
    elif stock_quantity < target_stock:
        return max_stock - stock_quantity
    else:
        return max(0, 50 - stock_quantity)


# ============================================================================
# API Endpoints
# ============================================================================


@router.get("/statistics")
async def get_inventory_statistics(
    account_type: str | None = Query(
        None, description="Filter by account type: main or fbe"
    ),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
) -> dict[str, Any]:
    """
    Get inventory statistics for eMAG products.

    Returns:
        - Total products
        - Low stock count
        - Out of stock count
        - Average stock level
        - Statistics by account type

    Caching: Results are cached for 5 minutes to reduce database load.
    """
    try:
        # Normalize account_type using utility function
        account_type = normalize_account_type(account_type)

        # Try to get from cache first
        if CACHE_AVAILABLE:
            cache = get_inventory_cache()
            cached_stats = await cache.get_statistics(account_type)
            if cached_stats:
                logger.debug(f"Cache hit for statistics (account_type={account_type})")
                return {"status": "success", "data": cached_stats, "cached": True}
        # Base query
        query = select(EmagProductV2)

        if account_type:
            query = query.where(EmagProductV2.account_type == account_type)

        # Get all products
        result = await db.execute(query)
        products = result.scalars().all()

        # Calculate statistics
        total_products = len(products)
        low_stock_count = 0
        out_of_stock_count = 0
        critical_count = 0
        total_stock = 0
        total_value = 0.0

        # Define thresholds
        CRITICAL_THRESHOLD = 5
        LOW_STOCK_THRESHOLD = 10

        for product in products:
            stock = product.stock_quantity or 0
            price = product.price or 0
            total_stock += stock
            total_value += stock * price

            if stock == 0:
                out_of_stock_count += 1
            elif stock <= CRITICAL_THRESHOLD:
                critical_count += 1
            elif stock <= LOW_STOCK_THRESHOLD:
                low_stock_count += 1

        avg_stock = total_stock / total_products if total_products > 0 else 0
        in_stock_count = total_products - out_of_stock_count
        stock_health = (
            (in_stock_count / total_products * 100) if total_products > 0 else 0
        )

        # Get statistics by account type
        main_query = select(func.count(EmagProductV2.id)).where(
            EmagProductV2.account_type == "main"
        )
        main_result = await db.execute(main_query)
        main_count = main_result.scalar() or 0

        fbe_query = select(func.count(EmagProductV2.id)).where(
            EmagProductV2.account_type == "fbe"
        )
        fbe_result = await db.execute(fbe_query)
        fbe_count = fbe_result.scalar() or 0

        # Calculate needs_reorder (products with stock <= LOW_STOCK_THRESHOLD)
        needs_reorder = out_of_stock_count + critical_count + low_stock_count

        statistics_data = {
            "total_items": total_products,
            "out_of_stock": out_of_stock_count,
            "critical": critical_count,
            "low_stock": low_stock_count,
            "in_stock": in_stock_count,
            "needs_reorder": needs_reorder,
            "total_value": round(total_value, 2),
            "stock_health_percentage": round(stock_health, 2),
            "average_stock_level": round(avg_stock, 2),
            "by_account": {
                "main": main_count,
                "fbe": fbe_count,
            },
            "thresholds": {
                "critical": CRITICAL_THRESHOLD,
                "low_stock": LOW_STOCK_THRESHOLD,
            },
        }

        # Cache the results
        if CACHE_AVAILABLE:
            cache = get_inventory_cache()
            await cache.set_statistics(statistics_data, account_type)
            logger.debug(f"Cached statistics (account_type={account_type})")

        return {"status": "success", "data": statistics_data, "cached": False}

    except Exception as e:
        logger.error(f"Error fetching inventory statistics: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch inventory statistics: {str(e)}",
        )


@router.get("/low-stock")
async def get_low_stock_products(
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(20, ge=1, le=100, description="Maximum number of records"),
    account_type: str | None = Query(
        None, description="Filter by account type: main or fbe"
    ),
    group_by_sku: bool = Query(
        True, description="Group products by SKU across accounts"
    ),
    threshold: int = Query(
        10, ge=0, description="Stock level threshold for low stock alert"
    ),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
) -> dict[str, Any]:
    """
    Get products with low stock levels.

    Args:
        skip: Pagination offset
        limit: Maximum number of results
        account_type: Filter by account (main/fbe)
        group_by_sku: Group products by SKU
        threshold: Stock level threshold

    Returns:
        List of products with low stock
    """
    try:
        # Normalize account_type using utility function
        account_type = normalize_account_type(account_type)

        # Build optimized query with filters
        filters = [
            or_(
                EmagProductV2.stock_quantity <= threshold,
                EmagProductV2.stock_quantity.is_(None),
            )
        ]

        if account_type:
            filters.append(EmagProductV2.account_type == account_type)

        # Base query for low stock products
        query = select(EmagProductV2).where(and_(*filters))

        # Order by stock level (lowest first)
        query = query.order_by(EmagProductV2.stock_quantity.asc().nulls_first())

        # Get total count efficiently
        count_query = select(func.count(EmagProductV2.id)).where(and_(*filters))
        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination
        query = query.offset(skip).limit(limit)

        # Execute query
        result = await db.execute(query)
        products = result.scalars().all()

        # Format response
        products_data = []

        if group_by_sku:
            # Group by SKU and aggregate data
            sku_groups = {}
            for product in products:
                sku = product.part_number_key or product.sku or str(product.id)

                if sku not in sku_groups:
                    stock = product.stock_quantity or 0
                    # Determine stock status
                    if stock == 0:
                        stock_status = "out_of_stock"
                    elif stock <= 5:
                        stock_status = "critical"
                    elif stock <= 10:
                        stock_status = "low_stock"
                    else:
                        stock_status = "in_stock"

                    sku_groups[sku] = {
                        "id": str(product.id),
                        "part_number_key": sku,
                        "sku": sku,
                        "name": product.name,
                        "brand": product.brand,
                        "category_name": product.emag_category_name,
                        "accounts": [],
                        "total_stock": 0,
                        "min_stock": stock,
                        "stock_status": stock_status,
                        "price": product.price or 0,
                        "currency": product.currency or "RON",
                        "account_type": "MULTI",
                    }

                stock = product.stock_quantity or 0
                sku_groups[sku]["accounts"].append(
                    {
                        "account_type": product.account_type.upper()
                        if product.account_type
                        else "MAIN",
                        "product_id": str(product.id),
                        "emag_id": product.emag_id,
                        "stock": stock,
                        "sale_price": product.price,
                        "status": product.status,
                    }
                )

                sku_groups[sku]["total_stock"] += stock
                sku_groups[sku]["min_stock"] = min(sku_groups[sku]["min_stock"], stock)

                # Update main_stock and fbe_stock
                if product.account_type and product.account_type.lower() == "main":
                    sku_groups[sku]["main_stock"] = stock
                elif product.account_type and product.account_type.lower() == "fbe":
                    sku_groups[sku]["fbe_stock"] = stock

            # Add calculated fields
            for _sku, group in sku_groups.items():
                total = group["total_stock"]
                # Update stock_status based on total
                if total == 0:
                    group["stock_status"] = "out_of_stock"
                elif total <= 5:
                    group["stock_status"] = "critical"
                elif total <= 10:
                    group["stock_status"] = "low_stock"
                else:
                    group["stock_status"] = "in_stock"

                # Calculate reorder quantity
                group["reorder_quantity"] = max(0, 20 - total)
                group["stock_quantity"] = total

            products_data = list(sku_groups.values())
        else:
            # Return individual products
            for product in products:
                stock = product.stock_quantity or 0
                stock_status = calculate_stock_status(stock)
                reorder_qty = calculate_reorder_quantity(stock)

                products_data.append(
                    {
                        "id": str(product.id),
                        "emag_id": product.emag_id,
                        "name": product.name,
                        "part_number_key": product.part_number_key or product.sku,
                        "sku": product.sku,
                        "account_type": product.account_type.upper()
                        if product.account_type
                        else "MAIN",
                        "stock_quantity": stock,
                        "price": product.price or 0,
                        "currency": product.currency or "RON",
                        "stock_status": stock_status,
                        "reorder_quantity": reorder_qty,
                        "sale_price": product.price,
                        "status": product.status,
                        "brand": product.brand,
                        "category_name": product.emag_category_name,
                        "alert_level": "critical" if stock == 0 else "warning",
                    }
                )

        return {
            "status": "success",
            "data": {
                "products": products_data,
                "pagination": {
                    "total": total,
                    "skip": skip,
                    "limit": limit,
                    "has_more": skip + limit < total,
                },
                "threshold": threshold,
                "grouped_by_sku": group_by_sku,
            },
        }

    except Exception as e:
        logger.error(f"Error fetching low stock products: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch low stock products: {str(e)}",
        )


@router.get("/stock-alerts")
async def get_stock_alerts(
    severity: str | None = Query(
        None, description="Filter by severity: critical, warning, info"
    ),
    account_type: str | None = Query(
        None, description="Filter by account type: main or fbe"
    ),
    limit: int = Query(50, ge=1, le=200, description="Maximum number of alerts"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
) -> dict[str, Any]:
    """
    Get stock alerts based on predefined thresholds.

    Severity levels:
    - critical: Out of stock (stock = 0)
    - warning: Low stock (stock <= 10)
    - info: Medium stock (stock <= 50)
    """
    try:
        # Normalize account_type using utility function
        account_type = normalize_account_type(account_type)

        # Define thresholds
        CRITICAL_THRESHOLD = 0
        WARNING_THRESHOLD = 10
        INFO_THRESHOLD = 50

        # Build query based on severity
        if severity == "critical":
            query = select(EmagProductV2).where(
                or_(
                    EmagProductV2.stock_quantity == CRITICAL_THRESHOLD,
                    EmagProductV2.stock_quantity.is_(None),
                )
            )
        elif severity == "warning":
            query = select(EmagProductV2).where(
                and_(
                    EmagProductV2.stock_quantity > CRITICAL_THRESHOLD,
                    EmagProductV2.stock_quantity <= WARNING_THRESHOLD,
                )
            )
        elif severity == "info":
            query = select(EmagProductV2).where(
                and_(
                    EmagProductV2.stock_quantity > WARNING_THRESHOLD,
                    EmagProductV2.stock_quantity <= INFO_THRESHOLD,
                )
            )
        else:
            # All alerts
            query = select(EmagProductV2).where(
                or_(
                    EmagProductV2.stock_quantity <= INFO_THRESHOLD,
                    EmagProductV2.stock_quantity.is_(None),
                )
            )

        if account_type:
            query = query.where(EmagProductV2.account_type == account_type)

        # Order by stock level
        query = query.order_by(EmagProductV2.stock_quantity.asc().nulls_first()).limit(
            limit
        )

        # Execute query
        result = await db.execute(query)
        products = result.scalars().all()

        # Format alerts
        alerts = []
        for product in products:
            stock = product.stock_quantity or 0

            # Determine severity
            if stock == 0:
                alert_severity = "critical"
                message = f"Out of stock: {product.name}"
            elif stock <= WARNING_THRESHOLD:
                alert_severity = "warning"
                message = f"Low stock ({stock} units): {product.name}"
            else:
                alert_severity = "info"
                message = f"Medium stock ({stock} units): {product.name}"

            alerts.append(
                {
                    "severity": alert_severity,
                    "message": message,
                    "product": {
                        "id": product.id,
                        "emag_id": product.emag_id,
                        "name": product.name,
                        "sku": product.sku,
                        "account_type": product.account_type,
                        "stock": stock,
                        "sale_price": product.price,
                    },
                    "timestamp": product.updated_at.isoformat()
                    if product.updated_at
                    else None,
                }
            )

        return {
            "status": "success",
            "data": {
                "alerts": alerts,
                "count": len(alerts),
                "severity_filter": severity,
                "thresholds": {
                    "critical": CRITICAL_THRESHOLD,
                    "warning": WARNING_THRESHOLD,
                    "info": INFO_THRESHOLD,
                },
            },
        }

    except Exception as e:
        logger.error(f"Error fetching stock alerts: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch stock alerts: {str(e)}",
        )


@router.get("/export/low-stock-excel")
async def export_low_stock_to_excel(
    account_type: str | None = Query(
        None, description="Filter by account: MAIN or FBE"
    ),
    stock_status: str | None = Query(
        None, description="Filter by status: critical, low_stock, out_of_stock"
    ),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Export low stock products to Excel file.

    Creates a formatted Excel file with:
    - Product details
    - Stock levels
    - Reorder recommendations
    - Pricing information
    """

    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export not available. Please install openpyxl package.",
        )

    try:
        # Normalize account_type using utility function
        account_type = normalize_account_type(account_type)

        # Build query for low stock products
        LOW_STOCK_THRESHOLD = 20

        query = select(EmagProductV2).where(
            or_(
                EmagProductV2.stock_quantity <= LOW_STOCK_THRESHOLD,
                EmagProductV2.stock_quantity.is_(None),
            )
        )

        if account_type:
            query = query.where(EmagProductV2.account_type == account_type)

        # Filter by status
        if stock_status == "out_of_stock":
            query = query.where(
                or_(
                    EmagProductV2.stock_quantity == 0,
                    EmagProductV2.stock_quantity.is_(None),
                )
            )
        elif stock_status == "critical":
            query = query.where(
                and_(
                    EmagProductV2.stock_quantity > 0, EmagProductV2.stock_quantity <= 10
                )
            )
        elif stock_status == "low_stock":
            query = query.where(
                and_(
                    EmagProductV2.stock_quantity > 10,
                    EmagProductV2.stock_quantity <= LOW_STOCK_THRESHOLD,
                )
            )

        # Order by stock level (lowest first)
        query = query.order_by(EmagProductV2.stock_quantity.asc().nulls_first())

        # Execute query
        result = await db.execute(query)
        products = result.scalars().all()

        if not products:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No low stock products found",
            )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Low Stock Products"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        critical_fill = PatternFill(
            start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
        )
        low_stock_fill = PatternFill(
            start_color="FFD93D", end_color="FFD93D", fill_type="solid"
        )

        # Define headers
        headers = [
            "Part Number",
            "Product Name",
            "Account Type",
            "Current Stock",
            "Status",
            "Reorder Qty",
            "Unit Price",
            "Total Cost",
            "Currency",
            "Brand",
            "Category",
            "EAN",
            "Last Updated",
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        row_num = 2
        total_reorder_cost = 0

        for product in products:
            stock = product.stock_quantity or 0

            # Determine stock status
            if stock == 0:
                stock_status = "OUT OF STOCK"
                fill_color = critical_fill
            elif stock <= 10:
                stock_status = "CRITICAL"
                fill_color = critical_fill
            elif stock <= 20:
                stock_status = "LOW STOCK"
                fill_color = low_stock_fill
            else:
                stock_status = "IN STOCK"
                fill_color = None

            # Calculate reorder quantity (target: 20 units)
            reorder_qty = max(0, 20 - stock)
            unit_price = product.price or 0
            total_cost = unit_price * reorder_qty
            total_reorder_cost += total_cost

            # Row data
            row_data = [
                product.part_number_key or product.sku or "",
                product.name or "",
                (product.account_type or "MAIN").upper(),
                stock,
                stock_status,
                reorder_qty,
                unit_price,
                total_cost,
                product.currency or "RON",
                product.brand or "",
                product.emag_category_name or "",
                product.ean or "",
                product.updated_at.strftime("%Y-%m-%d %H:%M")
                if product.updated_at
                else "",
            ]

            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

                # Apply conditional formatting
                if fill_color and col_num <= 5:
                    cell.fill = fill_color

            row_num += 1

        # Add summary section
        row_num += 2
        summary_cell = ws.cell(row=row_num, column=1)
        summary_cell.value = "SUMMARY"
        summary_cell.font = Font(bold=True, size=14)

        row_num += 1
        ws.cell(row=row_num, column=1).value = "Total Products:"
        ws.cell(row=row_num, column=2).value = len(products)
        ws.cell(row=row_num, column=1).font = Font(bold=True)

        row_num += 1
        ws.cell(row=row_num, column=1).value = "Total Reorder Cost:"
        ws.cell(row=row_num, column=2).value = f"{total_reorder_cost:.2f} RON"
        ws.cell(row=row_num, column=1).font = Font(bold=True)

        row_num += 1
        ws.cell(row=row_num, column=1).value = "Generated:"
        ws.cell(row=row_num, column=2).value = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        # Set column widths
        column_widths = {
            "A": 20,  # Part Number
            "B": 50,  # Product Name
            "C": 15,  # Account Type
            "D": 15,  # Current Stock
            "E": 15,  # Status
            "F": 15,  # Reorder Qty
            "G": 12,  # Unit Price
            "H": 15,  # Total Cost
            "I": 10,  # Currency
            "J": 20,  # Brand
            "K": 30,  # Category
            "L": 15,  # EAN
            "M": 20,  # Last Updated
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"low_stock_products_{timestamp}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting low stock to Excel: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export low stock products: {str(e)}",
        )


@router.get("/search")
async def search_emag_products(
    query: str = Query(
        ..., min_length=2, description="Search by SKU, part_number_key, or name"
    ),
    limit: int = Query(20, ge=1, le=100, description="Maximum number of results"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
) -> dict[str, Any]:
    """
    Search for eMAG products by SKU, part_number_key, or name.

    Returns products from both MAIN and FBE accounts with stock breakdown.
    Groups results by SKU to show combined stock across accounts.

    Args:
        query: Search term (minimum 2 characters)
        limit: Maximum number of results (default: 20, max: 100)

    Returns:
        List of products with stock breakdown by account
    """
    try:
        # Try to get from cache first
        if CACHE_AVAILABLE:
            cache = get_inventory_cache()
            cached_results = await cache.get_search_results(query, limit)
            if cached_results:
                logger.debug(f"Cache hit for search query: {query}")
                cached_results["cached"] = True
                return cached_results

        # Search for products
        search_query = (
            select(EmagProductV2)
            .where(
                or_(
                    EmagProductV2.sku.ilike(f"%{query}%"),
                    EmagProductV2.part_number_key.ilike(f"%{query}%"),
                    EmagProductV2.name.ilike(f"%{query}%"),
                )
            )
            .order_by(EmagProductV2.updated_at.desc())
            .limit(limit * 2)
        )  # Get more to account for grouping

        result = await db.execute(search_query)
        products = result.scalars().all()

        if not products:
            return {
                "status": "success",
                "data": {
                    "products": [],
                    "total": 0,
                    "query": query,
                },
            }

        # Group by SKU to show stock breakdown
        products_by_sku = {}
        for product in products:
            sku = product.sku or product.part_number_key
            if not sku:
                continue

            if sku not in products_by_sku:
                products_by_sku[sku] = {
                    "id": str(product.id),
                    "sku": sku,
                    "part_number_key": product.part_number_key,
                    "name": product.name,
                    "main_stock": 0,
                    "fbe_stock": 0,
                    "total_stock": 0,
                    "price": product.price or 0,
                    "currency": product.currency or "RON",
                    "brand": product.brand,
                    "category_name": product.emag_category_name,
                    "ean": product.ean,
                    "accounts": [],
                }

            stock = product.stock_quantity or 0
            account_type = (product.account_type or "MAIN").upper()

            # Add account details
            products_by_sku[sku]["accounts"].append(
                {
                    "account_type": account_type,
                    "product_id": str(product.id),
                    "emag_id": product.emag_id,
                    "stock": stock,
                    "price": product.price,
                    "status": product.status,
                }
            )

            # Update stock breakdown
            if account_type == "MAIN":
                products_by_sku[sku]["main_stock"] = stock
            elif account_type == "FBE":
                products_by_sku[sku]["fbe_stock"] = stock

            products_by_sku[sku]["total_stock"] += stock

        # Calculate stock status for each product
        for _sku, product_data in products_by_sku.items():
            total_stock = product_data["total_stock"]
            product_data["stock_status"] = calculate_stock_status(total_stock)
            product_data["reorder_quantity"] = calculate_reorder_quantity(total_stock)

        # Limit results after grouping
        products_list = list(products_by_sku.values())[:limit]

        result_data = {
            "status": "success",
            "data": {
                "products": products_list,
                "total": len(products_list),
                "query": query,
            },
            "cached": False,
        }

        # Cache the results
        if CACHE_AVAILABLE:
            cache = get_inventory_cache()
            await cache.set_search_results(query, result_data, limit)
            logger.debug(f"Cached search results for query: {query}")

        return result_data

    except Exception as e:
        logger.error(f"Error searching products: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to search products: {str(e)}",
        )
