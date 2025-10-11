"""
Inventory Management API with Low Stock Alerts and Excel Export.

Features:
- Low stock product identification
- Excel export for supplier orders
- Stock level monitoring
- Reorder point management
"""

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.db import get_db
from app.models.inventory import InventoryItem, Warehouse
from app.models.product import Product
from app.models.supplier import SupplierProduct
from app.models.user import User
from app.security.jwt import get_current_user

router = APIRouter(prefix="/inventory", tags=["inventory-management"])


# ============================================================================
# Helper Functions
# ============================================================================


def calculate_stock_status(item: InventoryItem) -> str:
    """Calculate stock status based on quantity and thresholds."""
    available = item.quantity - item.reserved_quantity

    if available <= 0:
        return "out_of_stock"
    elif available <= item.minimum_stock:
        return "critical"
    elif available <= item.reorder_point:
        return "low_stock"
    elif item.maximum_stock and available >= item.maximum_stock:
        return "overstock"
    else:
        return "in_stock"


def calculate_reorder_quantity(item: InventoryItem) -> int:
    """Calculate recommended reorder quantity."""
    available = item.quantity - item.reserved_quantity

    if item.maximum_stock:
        # Reorder to maximum stock
        return max(0, item.maximum_stock - available)
    elif item.reorder_point > 0:
        # Reorder to double the reorder point
        return max(0, (item.reorder_point * 2) - available)
    else:
        # Default: reorder to minimum stock * 3
        return max(0, (item.minimum_stock * 3) - available)


# ============================================================================
# Inventory Endpoints
# ============================================================================


@router.get("/low-stock")
async def get_low_stock_products(
    warehouse_id: int | None = Query(None, description="Filter by warehouse"),
    status: str | None = Query(
        None, description="Filter by status: critical, low_stock, out_of_stock"
    ),
    include_inactive: bool = Query(False, description="Include inactive products"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get products with low stock levels.

    Returns products that need reordering based on:
    - Out of stock (quantity <= 0)
    - Critical (available <= minimum_stock)
    - Low stock (available <= reorder_point)
    """

    # Build query
    query = (
        select(InventoryItem, Product, Warehouse)
        .join(Product, InventoryItem.product_id == Product.id)
        .join(Warehouse, InventoryItem.warehouse_id == Warehouse.id)
        .where(InventoryItem.is_active.is_(True))
    )

    # Filter by warehouse
    if warehouse_id:
        query = query.where(InventoryItem.warehouse_id == warehouse_id)

    # Filter by product active status
    if not include_inactive:
        query = query.where(Product.is_active.is_(True))

    # Filter by stock status
    if status == "out_of_stock":
        query = query.where(InventoryItem.quantity <= 0)
    elif status == "critical":
        query = query.where(
            and_(
                InventoryItem.quantity > 0,
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                <= InventoryItem.minimum_stock,
            )
        )
    elif status == "low_stock":
        query = query.where(
            and_(
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                > InventoryItem.minimum_stock,
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                <= InventoryItem.reorder_point,
            )
        )
    else:
        # Default: show all products that need attention
        query = query.where(
            or_(
                InventoryItem.quantity <= 0,
                (InventoryItem.quantity - InventoryItem.reserved_quantity)
                <= InventoryItem.reorder_point,
            )
        )

    # Order by urgency (out of stock first, then by available quantity)
    query = query.order_by(
        InventoryItem.quantity.asc(),
        (InventoryItem.quantity - InventoryItem.reserved_quantity).asc(),
    )

    # Get total count
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Apply pagination
    query = query.offset(skip).limit(limit)

    # Execute query
    result = await db.execute(query)
    items = result.all()

    # Format response
    products_data = []
    for inventory_item, product, warehouse in items:
        available_qty = inventory_item.quantity - inventory_item.reserved_quantity
        stock_status = calculate_stock_status(inventory_item)
        reorder_qty = calculate_reorder_quantity(inventory_item)

        products_data.append(
            {
                "inventory_item_id": inventory_item.id,
                "product_id": product.id,
                "sku": product.sku,
                "name": product.name,
                "chinese_name": product.chinese_name,
                "warehouse_id": warehouse.id,
                "warehouse_name": warehouse.name,
                "warehouse_code": warehouse.code,
                "quantity": inventory_item.quantity,
                "reserved_quantity": inventory_item.reserved_quantity,
                "available_quantity": available_qty,
                "minimum_stock": inventory_item.minimum_stock,
                "reorder_point": inventory_item.reorder_point,
                "maximum_stock": inventory_item.maximum_stock,
                "unit_cost": inventory_item.unit_cost,
                "stock_status": stock_status,
                "reorder_quantity": reorder_qty,
                "location": inventory_item.location,
                "batch_number": inventory_item.batch_number,
                "base_price": product.base_price,
                "currency": product.currency,
                "is_discontinued": product.is_discontinued,
            }
        )

    return {
        "status": "success",
        "data": {
            "products": products_data,
            "pagination": {
                "total": total,
                "skip": skip,
                "limit": limit,
                "has_more": skip + limit < total,
            },
            "summary": {
                "total_low_stock": total,
                "out_of_stock": sum(
                    1 for p in products_data if p["stock_status"] == "out_of_stock"
                ),
                "critical": sum(
                    1 for p in products_data if p["stock_status"] == "critical"
                ),
                "low_stock": sum(
                    1 for p in products_data if p["stock_status"] == "low_stock"
                ),
            },
        },
    }


@router.get("/export/low-stock-excel")
async def export_low_stock_to_excel(
    warehouse_id: int | None = Query(None, description="Filter by warehouse"),
    status: str | None = Query(None, description="Filter by status"),
    include_supplier_info: bool = Query(
        True, description="Include supplier information"
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Export low stock products to Excel file for supplier orders.

    The Excel file includes:
    - Product details (SKU, Name, Description)
    - Current stock levels
    - Recommended reorder quantities
    - Supplier information (if available)
    - Pricing information
    """

    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl: pip install openpyxl",
        )

    # Get low stock products (no pagination for export)
    query = (
        select(InventoryItem, Product, Warehouse)
        .join(Product, InventoryItem.product_id == Product.id)
        .join(Warehouse, InventoryItem.warehouse_id == Warehouse.id)
        .where(
            and_(
                InventoryItem.is_active.is_(True),
                Product.is_active.is_(True),
                or_(
                    InventoryItem.quantity <= 0,
                    (InventoryItem.quantity - InventoryItem.reserved_quantity)
                    <= InventoryItem.reorder_point,
                ),
            )
        )
    )

    if warehouse_id:
        query = query.where(InventoryItem.warehouse_id == warehouse_id)

    if status:
        if status == "out_of_stock":
            query = query.where(InventoryItem.quantity <= 0)
        elif status == "critical":
            query = query.where(
                and_(
                    InventoryItem.quantity > 0,
                    (InventoryItem.quantity - InventoryItem.reserved_quantity)
                    <= InventoryItem.minimum_stock,
                )
            )

    query = query.order_by(InventoryItem.quantity.asc())

    result = await db.execute(query)
    items = result.all()

    if not items:
        raise HTTPException(status_code=404, detail="No low stock products found")

    # Get supplier information if requested
    supplier_info = {}
    if include_supplier_info:
        product_ids = [item[1].id for item in items]
        supplier_query = (
            select(SupplierProduct)
            .options(selectinload(SupplierProduct.supplier))
            .where(SupplierProduct.local_product_id.in_(product_ids))
        )
        supplier_result = await db.execute(supplier_query)
        supplier_products = supplier_result.scalars().all()

        for sp in supplier_products:
            if sp.local_product_id not in supplier_info:
                supplier_info[sp.local_product_id] = []
            supplier_info[sp.local_product_id].append(
                {
                    "supplier_name": sp.supplier.name if sp.supplier else "Unknown",
                    "supplier_sku": sp.supplier_product_name,
                    "supplier_price": sp.supplier_price,
                    "supplier_url": sp.supplier_product_url,
                }
            )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Low Stock Products"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Critical status style
    critical_fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )
    critical_font = Font(color="FFFFFF", bold=True)

    # Low stock style
    low_stock_fill = PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    )

    # Headers
    headers = [
        "SKU",
        "Product Name",
        "Chinese Name",
        "Warehouse",
        "Current Stock",
        "Reserved",
        "Available",
        "Min Stock",
        "Reorder Point",
        "Status",
        "Reorder Qty",
        "Unit Cost",
        "Total Cost",
        "Supplier Name",
        "Supplier SKU",
        "Supplier Price",
        "Supplier URL",
        "Location",
        "Notes",
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    row_num = 2
    for inventory_item, product, warehouse in items:
        available_qty = inventory_item.quantity - inventory_item.reserved_quantity
        stock_status = calculate_stock_status(inventory_item)
        reorder_qty = calculate_reorder_quantity(inventory_item)

        # Get supplier info for this product
        suppliers = supplier_info.get(product.id, [])
        supplier_name = suppliers[0]["supplier_name"] if suppliers else ""
        supplier_sku = suppliers[0]["supplier_sku"] if suppliers else ""
        supplier_price = suppliers[0]["supplier_price"] if suppliers else None
        supplier_url = suppliers[0]["supplier_url"] if suppliers else ""

        # Calculate total cost
        total_cost = (inventory_item.unit_cost or 0) * reorder_qty

        # Write row data
        row_data = [
            product.sku,
            product.name,
            product.chinese_name or "",
            warehouse.name,
            inventory_item.quantity,
            inventory_item.reserved_quantity,
            available_qty,
            inventory_item.minimum_stock,
            inventory_item.reorder_point,
            stock_status.upper().replace("_", " "),
            reorder_qty,
            inventory_item.unit_cost or 0,
            total_cost,
            supplier_name,
            supplier_sku,
            supplier_price or 0,
            supplier_url,
            inventory_item.location or "",
            f"Order urgency: {stock_status}",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

            # Apply status-based styling
            if stock_status == "out_of_stock":
                cell.fill = critical_fill
                cell.font = critical_font
            elif stock_status == "critical":
                cell.fill = low_stock_fill

        row_num += 1

    # Add summary row
    row_num += 1
    summary_cell = ws.cell(row=row_num, column=1)
    summary_cell.value = "SUMMARY"
    summary_cell.font = Font(bold=True, size=12)

    row_num += 1
    ws.cell(row=row_num, column=1).value = "Total Products:"
    ws.cell(row=row_num, column=2).value = len(items)
    ws.cell(row=row_num, column=1).font = Font(bold=True)

    row_num += 1
    ws.cell(row=row_num, column=1).value = "Total Reorder Cost:"
    total_reorder_cost = sum(
        (item[0].unit_cost or 0) * calculate_reorder_quantity(item[0]) for item in items
    )
    ws.cell(row=row_num, column=2).value = f"{total_reorder_cost:.2f} RON"
    ws.cell(row=row_num, column=1).font = Font(bold=True)

    row_num += 1
    ws.cell(row=row_num, column=1).value = "Generated:"
    ws.cell(row=row_num, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Adjust column widths
    column_widths = {
        "A": 15,  # SKU
        "B": 40,  # Product Name
        "C": 30,  # Chinese Name
        "D": 20,  # Warehouse
        "E": 12,  # Current Stock
        "F": 12,  # Reserved
        "G": 12,  # Available
        "H": 12,  # Min Stock
        "I": 12,  # Reorder Point
        "J": 15,  # Status
        "K": 12,  # Reorder Qty
        "L": 12,  # Unit Cost
        "M": 12,  # Total Cost
        "N": 25,  # Supplier Name
        "O": 20,  # Supplier SKU
        "P": 15,  # Supplier Price
        "Q": 40,  # Supplier URL
        "R": 15,  # Location
        "S": 30,  # Notes
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze first row
    ws.freeze_panes = "A2"

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"low_stock_products_{timestamp}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/statistics")
async def get_inventory_statistics(
    warehouse_id: int | None = Query(None, description="Filter by warehouse"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get inventory statistics and summary."""

    # Build base query
    query = select(InventoryItem).where(InventoryItem.is_active.is_(True))

    if warehouse_id:
        query = query.where(InventoryItem.warehouse_id == warehouse_id)

    result = await db.execute(query)
    items = result.scalars().all()

    # Calculate statistics
    total_items = len(items)
    out_of_stock = sum(1 for item in items if item.quantity <= 0)
    critical = sum(1 for item in items if calculate_stock_status(item) == "critical")
    low_stock = sum(1 for item in items if calculate_stock_status(item) == "low_stock")
    in_stock = sum(1 for item in items if calculate_stock_status(item) == "in_stock")
    overstock = sum(1 for item in items if calculate_stock_status(item) == "overstock")

    total_value = sum((item.quantity * (item.unit_cost or 0)) for item in items)
    total_reserved = sum(item.reserved_quantity for item in items)
    total_available = sum((item.quantity - item.reserved_quantity) for item in items)

    return {
        "status": "success",
        "data": {
            "total_items": total_items,
            "out_of_stock": out_of_stock,
            "critical": critical,
            "low_stock": low_stock,
            "in_stock": in_stock,
            "overstock": overstock,
            "needs_reorder": out_of_stock + critical + low_stock,
            "total_value": round(total_value, 2),
            "total_reserved": total_reserved,
            "total_available": total_available,
            "stock_health_percentage": round(
                ((in_stock + overstock) / total_items * 100) if total_items > 0 else 0,
                2,
            ),
        },
    }
