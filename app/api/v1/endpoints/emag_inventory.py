"""
eMAG Inventory Management - Direct integration with emag_products_v2.

This endpoint works directly with synchronized eMAG products to identify
low stock items and generate supplier orders.
"""

from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import String, and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.db import get_db
from app.models.emag_models import EmagProductV2
from app.models.user import User
from app.security.jwt import get_current_user

router = APIRouter(prefix="/emag-inventory", tags=["emag-inventory"])


def calculate_stock_status(stock_quantity: int, min_stock: int = 10, reorder_point: int = 20) -> str:
    """Calculate stock status based on quantity."""
    if stock_quantity <= 0:
        return "out_of_stock"
    elif stock_quantity <= min_stock:
        return "critical"
    elif stock_quantity <= reorder_point:
        return "low_stock"
    else:
        return "in_stock"


def calculate_reorder_quantity(stock_quantity: int, max_stock: int = 100) -> int:
    """Calculate recommended reorder quantity."""
    if stock_quantity <= 0:
        return max_stock
    elif stock_quantity < 20:
        return max_stock - stock_quantity
    else:
        return max(0, 50 - stock_quantity)


@router.get("/low-stock")
async def get_emag_low_stock_products(
    account_type: Optional[str] = Query(None, description="Filter by account: MAIN or FBE"),
    status: Optional[str] = Query(None, description="Filter by status: critical, low_stock, out_of_stock"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    group_by_sku: bool = Query(False, description="Group products by SKU and aggregate stock"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get eMAG products with low stock levels.
    
    Returns products from emag_products_v2 that need reordering based on:
    - Out of stock (stock_quantity <= 0)
    - Critical (stock_quantity <= 10)
    - Low stock (stock_quantity <= 20)
    
    New features:
    - group_by_sku: Aggregate stock across MAIN and FBE accounts
    - Shows stock breakdown per account
    """
    
    if group_by_sku:
        # Group by SKU and aggregate stock across accounts
        from sqlalchemy import case
        
        # Build having conditions
        having_conditions = []
        
        # Base condition: low stock products
        if not status or status == "all":
            having_conditions.append(
                or_(
                    func.sum(EmagProductV2.stock_quantity) <= 0,
                    func.sum(EmagProductV2.stock_quantity) <= 20
                )
            )
        elif status == "out_of_stock":
            having_conditions.append(func.sum(EmagProductV2.stock_quantity) <= 0)
        elif status == "critical":
            having_conditions.append(
                and_(
                    func.sum(EmagProductV2.stock_quantity) > 0,
                    func.sum(EmagProductV2.stock_quantity) <= 10
                )
            )
        elif status == "low_stock":
            having_conditions.append(
                and_(
                    func.sum(EmagProductV2.stock_quantity) > 10,
                    func.sum(EmagProductV2.stock_quantity) <= 20
                )
            )
        
        # Build where conditions
        where_conditions = [EmagProductV2.is_active.is_(True)]
        
        # Filter by account type if specified
        if account_type:
            where_conditions.append(func.lower(EmagProductV2.account_type) == account_type.lower())
        
        query = select(
            func.min(func.cast(EmagProductV2.id, String)).label('id'),
            EmagProductV2.sku,
            EmagProductV2.part_number_key,
            func.max(EmagProductV2.name).label('name'),
            func.sum(EmagProductV2.stock_quantity).label('total_stock'),
            func.sum(case((func.lower(EmagProductV2.account_type) == 'main', EmagProductV2.stock_quantity), else_=0)).label('main_stock'),
            func.sum(case((func.lower(EmagProductV2.account_type) == 'fbe', EmagProductV2.stock_quantity), else_=0)).label('fbe_stock'),
            func.max(EmagProductV2.price).label('price'),
            func.max(EmagProductV2.currency).label('currency'),
            func.max(EmagProductV2.best_offer_sale_price).label('sale_price'),
            func.max(EmagProductV2.best_offer_recommended_price).label('recommended_price'),
            func.max(EmagProductV2.vat_id).label('vat_id'),
            func.max(EmagProductV2.brand).label('brand'),
            func.max(EmagProductV2.emag_category_name).label('category_name'),
        ).where(
            and_(*where_conditions)
        ).group_by(
            EmagProductV2.sku,
            EmagProductV2.part_number_key
        )
        
        # Apply having conditions
        if having_conditions:
            query = query.having(and_(*having_conditions))
        
        # Order by urgency
        query = query.order_by(func.sum(EmagProductV2.stock_quantity).asc())
        
        # Get total count
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0
        
        # Apply pagination
        query = query.offset(skip).limit(limit)
        
        # Execute query
        result = await db.execute(query)
        rows = result.all()
        
        # Format response
        products_data = []
        for row in rows:
            total_stock = int(row.total_stock or 0)
            main_stock = int(row.main_stock or 0)
            fbe_stock = int(row.fbe_stock or 0)
            stock_status = calculate_stock_status(total_stock)
            reorder_qty = calculate_reorder_quantity(total_stock)
            
            products_data.append({
                "id": str(row.id),
                "part_number_key": row.part_number_key,
                "name": row.name,
                "account_type": "BOTH" if main_stock > 0 and fbe_stock > 0 else ("MAIN" if main_stock > 0 else "FBE"),
                "stock_quantity": total_stock,
                "main_stock": main_stock,
                "fbe_stock": fbe_stock,
                "price": row.price or 0,
                "currency": row.currency or "RON",
                "stock_status": stock_status,
                "reorder_quantity": reorder_qty,
                "sale_price": row.sale_price,
                "recommended_price": row.recommended_price,
                "vat_id": row.vat_id,
                "ean": None,
                "brand": row.brand,
                "category_name": row.category_name,
            })
    else:
        # Original behavior - show products separately per account
        query = select(EmagProductV2).where(
            and_(
                EmagProductV2.is_active.is_(True),
                or_(
                    EmagProductV2.stock_quantity <= 0,
                    EmagProductV2.stock_quantity <= 20
                )
            )
        )
        
        # Filter by account type
        if account_type:
            query = query.where(func.lower(EmagProductV2.account_type) == account_type.lower())
        
        # Filter by stock status
        if status == "out_of_stock":
            query = query.where(EmagProductV2.stock_quantity <= 0)
        elif status == "critical":
            query = query.where(
                and_(
                    EmagProductV2.stock_quantity > 0,
                    EmagProductV2.stock_quantity <= 10
                )
            )
        elif status == "low_stock":
            query = query.where(
                and_(
                    EmagProductV2.stock_quantity > 10,
                    EmagProductV2.stock_quantity <= 20
                )
            )
        
        # Order by urgency
        query = query.order_by(EmagProductV2.stock_quantity.asc())
        
        # Get total count
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0
        
        # Apply pagination
        query = query.offset(skip).limit(limit)
        
        # Execute query
        result = await db.execute(query)
        products = result.scalars().all()
        
        # Format response with stock breakdown
        products_data = []
        for product in products:
            stock_status = calculate_stock_status(product.stock_quantity or 0)
            reorder_qty = calculate_reorder_quantity(product.stock_quantity or 0)
            
            # Get stock from other account for the same SKU
            other_account = "FBE" if product.account_type == "MAIN" else "MAIN"
            other_stock_query = select(EmagProductV2.stock_quantity).where(
                and_(
                    EmagProductV2.sku == product.sku,
                    EmagProductV2.account_type == other_account,
                    EmagProductV2.is_active.is_(True)
                )
            )
            other_stock_result = await db.execute(other_stock_query)
            other_stock = other_stock_result.scalar() or 0
            
            products_data.append({
                "id": str(product.id),
                "part_number_key": product.part_number_key,
                "name": product.name,
                "account_type": product.account_type,
                "stock_quantity": product.stock_quantity or 0,
                "main_stock": product.stock_quantity if product.account_type == "MAIN" else other_stock,
                "fbe_stock": product.stock_quantity if product.account_type == "FBE" else other_stock,
                "price": product.price or 0,
                "currency": product.currency or "RON",
                "stock_status": stock_status,
                "reorder_quantity": reorder_qty,
                "sale_price": product.best_offer_sale_price,
                "recommended_price": product.best_offer_recommended_price,
                "vat_id": product.vat_id,
                "ean": product.ean,
                "brand": product.brand,
                "category_name": product.emag_category_name,
            })
    
    # Calculate summary
    out_of_stock = sum(1 for p in products_data if p["stock_status"] == "out_of_stock")
    critical = sum(1 for p in products_data if p["stock_status"] == "critical")
    low_stock = sum(1 for p in products_data if p["stock_status"] == "low_stock")
    
    return {
        "status": "success",
        "data": {
            "products": products_data,
            "pagination": {
                "total": total,
                "skip": skip,
                "limit": limit,
                "has_more": skip + limit < total,
            },
            "summary": {
                "total_low_stock": total,
                "out_of_stock": out_of_stock,
                "critical": critical,
                "low_stock": low_stock,
            }
        }
    }


@router.get("/statistics")
async def get_emag_inventory_statistics(
    account_type: Optional[str] = Query(None, description="Filter by account: MAIN or FBE"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get eMAG inventory statistics."""
    
    # Build base query
    query = select(EmagProductV2).where(EmagProductV2.is_active.is_(True))
    
    if account_type:
        query = query.where(func.lower(EmagProductV2.account_type) == account_type.lower())
    
    result = await db.execute(query)
    products = result.scalars().all()
    
    # Calculate statistics
    total_items = len(products)
    out_of_stock = sum(1 for p in products if (p.stock_quantity or 0) <= 0)
    critical = sum(1 for p in products if 0 < (p.stock_quantity or 0) <= 10)
    low_stock = sum(1 for p in products if 10 < (p.stock_quantity or 0) <= 20)
    in_stock = sum(1 for p in products if (p.stock_quantity or 0) > 20)
    
    total_value = sum((p.stock_quantity or 0) * (p.price or 0) for p in products)
    
    # Calculate by account
    main_products = sum(1 for p in products if p.account_type == "MAIN")
    fbe_products = sum(1 for p in products if p.account_type == "FBE")
    
    return {
        "status": "success",
        "data": {
            "total_items": total_items,
            "out_of_stock": out_of_stock,
            "critical": critical,
            "low_stock": low_stock,
            "in_stock": in_stock,
            "needs_reorder": out_of_stock + critical + low_stock,
            "total_value": round(total_value, 2),
            "stock_health_percentage": round((in_stock / total_items * 100) if total_items > 0 else 0, 2),
            "by_account": {
                "MAIN": main_products,
                "FBE": fbe_products,
            }
        }
    }


@router.get("/search")
async def search_emag_products(
    query: str = Query(..., description="Search by SKU, part_number_key, or name"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Search for eMAG products by SKU, part_number_key, or name.
    Returns products from both MAIN and FBE accounts with stock breakdown.
    """
    
    # Search for products
    search_query = select(EmagProductV2).where(
        and_(
            EmagProductV2.is_active.is_(True),
            or_(
                EmagProductV2.sku.ilike(f"%{query}%"),
                EmagProductV2.part_number_key.ilike(f"%{query}%"),
                EmagProductV2.name.ilike(f"%{query}%"),
            )
        )
    ).order_by(EmagProductV2.updated_at.desc()).limit(20)
    
    result = await db.execute(search_query)
    products = result.scalars().all()
    
    # Group by SKU to show stock breakdown
    products_by_sku = {}
    for product in products:
        sku = product.sku
        if sku not in products_by_sku:
            products_by_sku[sku] = {
                "id": str(product.id),
                "sku": sku,
                "part_number_key": product.part_number_key,
                "name": product.name,
                "main_stock": 0,
                "fbe_stock": 0,
                "total_stock": 0,
                "price": product.price or 0,
                "currency": product.currency or "RON",
                "brand": product.brand,
                "category_name": product.emag_category_name,
            }
        
        if product.account_type == "MAIN":
            products_by_sku[sku]["main_stock"] = product.stock_quantity or 0
        elif product.account_type == "FBE":
            products_by_sku[sku]["fbe_stock"] = product.stock_quantity or 0
        
        products_by_sku[sku]["total_stock"] = (
            products_by_sku[sku]["main_stock"] + products_by_sku[sku]["fbe_stock"]
        )
    
    return {
        "status": "success",
        "data": {
            "products": list(products_by_sku.values()),
            "total": len(products_by_sku),
        }
    }


@router.get("/export/low-stock-excel")
async def export_emag_low_stock_to_excel(
    account_type: Optional[str] = Query(None, description="Filter by account: MAIN or FBE"),
    status: Optional[str] = Query(None, description="Filter by status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export eMAG low stock products to Excel."""
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl"
        )
    
    # Get low stock products
    query = select(EmagProductV2).where(
        and_(
            EmagProductV2.is_active.is_(True),
            or_(
                EmagProductV2.stock_quantity <= 0,
                EmagProductV2.stock_quantity <= 20
            )
        )
    )
    
    if account_type:
        query = query.where(func.lower(EmagProductV2.account_type) == account_type.lower())
    
    if status == "out_of_stock":
        query = query.where(EmagProductV2.stock_quantity <= 0)
    elif status == "critical":
        query = query.where(
            and_(
                EmagProductV2.stock_quantity > 0,
                EmagProductV2.stock_quantity <= 10
            )
        )
    
    query = query.order_by(EmagProductV2.stock_quantity.asc())
    
    result = await db.execute(query)
    products = result.scalars().all()
    
    if not products:
        raise HTTPException(status_code=404, detail="No low stock products found")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "eMAG Low Stock"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    low_stock_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Headers
    headers = [
        "Part Number Key",
        "Product Name",
        "Account",
        "Current Stock",
        "Status",
        "Reorder Qty",
        "Price",
        "Sale Price",
        "Currency",
        "Total Cost",
        "EAN",
        "Brand",
        "Category",
        "VAT Rate",
        "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = 2
    for product in products:
        stock_qty = product.stock_quantity or 0
        stock_status = calculate_stock_status(stock_qty)
        reorder_qty = calculate_reorder_quantity(stock_qty)
        price = product.price or 0
        total_cost = price * reorder_qty
        
        row_data = [
            product.part_number_key or "",
            product.name or "",
            product.account_type or "",
            stock_qty,
            stock_status.upper().replace("_", " "),
            reorder_qty,
            price,
            product.best_offer_sale_price or 0,
            product.currency or "RON",
            total_cost,
            product.ean or "",
            product.brand or "",
            product.emag_category_name or "",
            product.vat_id or 0,
            f"Urgency: {stock_status}"
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            if stock_status == "out_of_stock":
                cell.fill = critical_fill
                cell.font = critical_font
            elif stock_status == "critical":
                cell.fill = low_stock_fill
        
        row_num += 1
    
    # Summary
    row_num += 1
    summary_cell = ws.cell(row=row_num, column=1)
    summary_cell.value = "SUMMARY"
    summary_cell.font = Font(bold=True, size=12)
    
    row_num += 1
    ws.cell(row=row_num, column=1).value = "Total Products:"
    ws.cell(row=row_num, column=2).value = len(products)
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    
    row_num += 1
    total_reorder_cost = sum((p.price or 0) * calculate_reorder_quantity(p.stock_quantity or 0) for p in products)
    ws.cell(row=row_num, column=1).value = "Total Reorder Cost:"
    ws.cell(row=row_num, column=2).value = f"{total_reorder_cost:.2f} RON"
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    
    row_num += 1
    ws.cell(row=row_num, column=1).value = "Generated:"
    ws.cell(row=row_num, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Column widths
    column_widths = {
        'A': 20, 'B': 50, 'C': 10, 'D': 12, 'E': 15,
        'F': 12, 'G': 12, 'H': 12, 'I': 10, 'J': 12,
        'K': 15, 'L': 20, 'M': 30, 'N': 10, 'O': 30
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"emag_low_stock_{timestamp}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
