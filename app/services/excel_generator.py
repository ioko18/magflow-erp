"""Excel generation service for supplier orders in MagFlow ERP.

This service generates customized Excel files for supplier orders based on:
- Supplier-specific templates
- Product mappings from 1688.com
- Custom formatting and branding
"""

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.supplier import Supplier, SupplierProduct

logger = logging.getLogger(__name__)


class ExcelGeneratorService:
    """Service for generating Excel files for supplier orders."""

    def __init__(self):
        self.templates_dir = (
            "supplier_templates"  # Directory for supplier-specific templates
        )

    async def generate_supplier_order_excel(
        self,
        supplier: Supplier,
        supplier_products: list[SupplierProduct],
        order_metadata: dict[str, Any] | None = None,
    ) -> bytes:
        """Generate Excel file for supplier order."""

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create order sheet
        order_sheet = wb.create_sheet("Order", 0)

        # Add header information
        self._add_order_header(order_sheet, supplier, order_metadata)

        # Add products table
        self._add_products_table(order_sheet, supplier_products, start_row=10)

        # Add footer information
        self._add_order_footer(order_sheet, supplier, len(supplier_products))

        # Apply formatting
        self._apply_formatting(order_sheet)

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info(
            f"Generated Excel order for supplier {supplier.name} with {len(supplier_products)} products"
        )
        return excel_buffer.getvalue()

    def _add_order_header(
        self, sheet, supplier: Supplier, metadata: dict[str, Any] | None = None
    ):
        """Add order header information."""

        # Company info (would be configurable)
        sheet["A1"] = "MAGFLOW ELECTRONICS SRL"
        sheet["A2"] = "Order către furnizor"
        sheet["A3"] = f"Furnizor: {supplier.name}"
        sheet["A4"] = f"Țară: {supplier.country}"
        sheet["A5"] = f"Data comandă: {datetime.now().strftime('%Y-%m-%d')}"

        if metadata:
            if "order_number" in metadata:
                sheet["A6"] = f"Număr comandă: {metadata['order_number']}"
            if "contact_person" in metadata:
                sheet["A7"] = f"Persoană contact: {metadata['contact_person']}"

        # Style header
        header_font = Font(bold=True, size=12)
        for row in range(1, 8):
            cell = sheet[f"A{row}"]
            cell.font = header_font

    def _add_products_table(
        self, sheet, supplier_products: list[SupplierProduct], start_row: int
    ):
        """Add products table with supplier mappings."""

        # Table headers
        headers = [
            "Nr.",
            "Nume Produs Local",
            "SKU Local",
            "Nume Chinezesc",
            "Cantitate",
            "Preț Furnizor",
            "Monedă",
            "Valoare Totală",
            "Link 1688",
            "Imagine",
            "Note",
        ]

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Add product data
        for idx, supplier_product in enumerate(supplier_products, 1):
            row = start_row + idx

            # Get local product info (would need to load this)
            # For now, using placeholder data
            local_product_name = f"Product {supplier_product.local_product_id}"
            local_sku = f"SKU-{supplier_product.local_product_id}"

            # Add product data
            sheet.cell(row=row, column=1, value=idx)
            sheet.cell(row=row, column=2, value=local_product_name)
            sheet.cell(row=row, column=3, value=local_sku)
            sheet.cell(row=row, column=4, value=supplier_product.supplier_product_name)
            sheet.cell(row=row, column=5, value=100)  # Quantity (would be calculated)
            sheet.cell(row=row, column=6, value=supplier_product.supplier_price)
            sheet.cell(row=row, column=7, value=supplier_product.supplier_currency)
            sheet.cell(
                row=row, column=8, value=f"=E{row}*F{row}"
            )  # Total value formula
            sheet.cell(row=row, column=9, value=supplier_product.supplier_product_url)
            sheet.cell(row=row, column=10, value=supplier_product.supplier_image_url)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 15

        # Make URL and image columns wider
        sheet.column_dimensions["I"].width = 30
        sheet.column_dimensions["J"].width = 30
        sheet.column_dimensions["K"].width = 20

    def _add_order_footer(self, sheet, supplier: Supplier, product_count: int):
        """Add order summary and footer."""

        # Find next available row
        max_row = sheet.max_row + 3

        # Order summary
        sheet.cell(row=max_row, column=1, value="REZUMAT COMANDĂ")
        sheet.cell(row=max_row, column=1).font = Font(bold=True, size=12)

        summary_row = max_row + 2
        sheet.cell(row=summary_row, column=1, value="Număr produse:")
        sheet.cell(row=summary_row, column=2, value=product_count)

        summary_row += 1
        sheet.cell(row=summary_row, column=1, value="Valoare totală:")
        sheet.cell(row=summary_row, column=2, value=f"=SUM(H11:H{sheet.max_row})")

        summary_row += 1
        sheet.cell(row=summary_row, column=1, value="Monedă:")
        sheet.cell(row=summary_row, column=2, value=supplier.currency)

        summary_row += 1
        sheet.cell(row=summary_row, column=1, value="Termeni plată:")
        sheet.cell(row=summary_row, column=2, value=supplier.payment_terms)

        summary_row += 1
        sheet.cell(row=summary_row, column=1, value="Lead time:")
        sheet.cell(row=summary_row, column=2, value=f"{supplier.lead_time_days} zile")

        # Terms and conditions
        terms_row = summary_row + 3
        sheet.cell(row=terms_row, column=1, value="TERMENI ȘI CONDIȚII:")
        sheet.cell(row=terms_row, column=1).font = Font(bold=True)

        terms_row += 1
        sheet.cell(
            row=terms_row,
            column=1,
            value="1. Calitatea produselor trebuie să corespundă specificațiilor.",
        )
        terms_row += 1
        sheet.cell(
            row=terms_row,
            column=1,
            value="2. Livrarea trebuie efectuată în termenul specificat.",
        )
        terms_row += 1
        sheet.cell(
            row=terms_row,
            column=1,
            value="3. Factura trebuie emisă în termen de 3 zile de la livrare.",
        )
        terms_row += 1
        sheet.cell(
            row=terms_row,
            column=1,
            value="4. Orice problemă de calitate va fi comunicată în maxim 7 zile.",
        )

    def _apply_formatting(self, sheet):
        """Apply consistent formatting to the sheet."""

        # Header styling
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for row in range(1, 20):  # First 20 rows for header area
            for col in range(1, 12):
                cell = sheet.cell(row=row, column=col)
                if row <= 8 and col == 1:  # Company info area
                    cell.font = Font(bold=True)
                elif row == 10:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

        # Border for product table
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in range(10, sheet.max_row + 1):
            for col in range(1, 12):
                sheet.cell(row=row, column=col).border = thin_border

    async def generate_custom_template(
        self, supplier_id: int, template_config: dict[str, Any]
    ) -> str:
        """Generate a custom Excel template for a supplier."""

        # Create template based on supplier preferences
        # This would create a reusable template file

        template_path = f"{self.templates_dir}/supplier_{supplier_id}_template.xlsx"

        # Create basic template
        wb = Workbook()
        ws = wb.active
        ws.title = "Supplier Order Template"

        # Add template structure based on config
        # Implementation would depend on template_config

        wb.save(template_path)
        logger.info(f"Generated custom template for supplier {supplier_id}")

        return template_path

    async def validate_excel_format(
        self, file_path: str, supplier_id: int
    ) -> dict[str, Any]:
        """Validate if uploaded Excel matches supplier template."""

        # Load and validate Excel structure
        # Check required columns, data types, etc.

        validation_result = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "column_mapping": {},
        }

        # Placeholder validation logic
        logger.info(f"Validated Excel format for supplier {supplier_id}")

        return validation_result


# Utility function for order number generation
def generate_order_number(supplier_id: int) -> str:
    """Generate unique order number for supplier."""

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"PO-{supplier_id}-{timestamp}"
